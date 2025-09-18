from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from schemas.ficha import Ficha


def get_image_path() -> Path:
    raiz = Path(__file__).parent.parent
    caminho_imagem = raiz / "assets/payer.png"
    try:
        return caminho_imagem
    except FileNotFoundError as e:
        print(f"Arquivo de imagem não encontrado. Detalhes: {e}")


def gera_planilha(dados: Ficha) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sitef"

    fonte_negrito = Font(bold=True)
    fonte_branca = Font(color="FFFFFF", bold=True)
    fonte_cinza_claro = Font(color="808080", bold=True)
    alinhamento_central = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    alinhamento_esquerda = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    preenchimento_cabecalho = PatternFill(
        start_color="171717", end_color="171717", fill_type="solid"
    )
    preenchimento_secoes = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )

    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 50

    ws["A1"].value = "PAYER"
    ws["A1"].font = fonte_branca
    ws["A1"].alignment = alinhamento_central
    ws["A1"].fill = preenchimento_cabecalho
    ws["A1"].border = borda_fina

    img = Image(get_image_path())
    img.width = 120
    img.height = 19
    ws.add_image(img, "A1")

    ws["B1"].value = "FICHA DE IMPLANTAÇÃO"
    ws["B1"].font = fonte_branca
    ws["B1"].alignment = alinhamento_central
    ws["B1"].fill = preenchimento_cabecalho
    ws["B1"].border = borda_fina

    linha_atual = 2
    for chave, valor in dados.model_dump().items():
        if chave == "account":
            ws.merge_cells(f"A{linha_atual}:B{linha_atual}")
            ws[f"A{linha_atual}"].value = "DADOS PAYER"
            ws[f"A{linha_atual}"].font = fonte_negrito
            ws[f"A{linha_atual}"].alignment = alinhamento_central
            ws[f"A{linha_atual}"].fill = preenchimento_secoes
            ws[f"A{linha_atual}"].border = borda_fina
            linha_atual += 1

        ws[f"A{linha_atual}"].value = chave.upper()
        ws[f"A{linha_atual}"].font = fonte_cinza_claro
        ws[f"A{linha_atual}"].border = borda_fina
        ws[f"A{linha_atual}"].alignment = alinhamento_central

        ws[f"B{linha_atual}"].value = valor
        ws[f"B{linha_atual}"].border = borda_fina
        ws[f"B{linha_atual}"].alignment = alinhamento_esquerda

        if linha_atual in range(13, 18):
            ws[f"B{linha_atual}"].alignment = alinhamento_central
            ws.column_dimensions["B"].width = 48

        linha_atual += 1

    return wb


def save(workbook: Workbook, ficha: Ficha) -> None:
    base_path = Path(r"G:\Drives compartilhados\FICHAS DE IMPLANTACAO")
    letra = ficha.razao_social[0]
    nome_arquivo = f"{ficha.store}.xlsx"

    if not base_path.exists():
        base_path = Path.home() / "Documents"
        workbook.save(base_path / nome_arquivo)
        print("Caminho do Drive não encontrado. Salvando em ~/Documentos.")
    else:
        pasta_letra = base_path / letra
        pasta_letra.mkdir(exist_ok=True)

        pasta_razao_social = pasta_letra / ficha.razao_social
        pasta_razao_social.mkdir(exist_ok=True)

        workbook.save(pasta_razao_social / nome_arquivo)
        print(f"Arquivo salvo em: {pasta_razao_social / nome_arquivo}")
